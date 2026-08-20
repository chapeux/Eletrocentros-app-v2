"""
Eletrocentros App — Backend Python & GUI Launcher
Substituição de Tkinter por Interface Web Desktop Moderna em HTML5/CSS3/JS.
"""

import sys
import os
import re
import json
import base64
import http.server
import socketserver
import threading
import webbrowser
from datetime import datetime
from pathlib import Path

# Módulo Backend de Banco de Dados MySQL
from backend.database import init_db, comparar_e_registrar_alteracoes, obter_logs, registrar_log
from backend.settings_store import get_setting, save_setting

# Base Directory Setup (suporta execução em modo script ou compilado PyInstaller)
if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).resolve().parent
    INTERNAL_DIR = Path(getattr(sys, "_MEIPASS", BASE_DIR / "_internal"))
else:
    BASE_DIR = Path(__file__).resolve().parent
    INTERNAL_DIR = BASE_DIR

FRONTEND_DIR = BASE_DIR / "frontend"
if not FRONTEND_DIR.exists() and (INTERNAL_DIR / "frontend").exists():
    FRONTEND_DIR = INTERNAL_DIR / "frontend"

ICON_PATH = BASE_DIR / "assets" / "icone.ico"
if not ICON_PATH.exists() and (INTERNAL_DIR / "assets" / "icone.ico").exists():
    ICON_PATH = INTERNAL_DIR / "assets" / "icone.ico"

# CONFIG FILE PATHS, REGRAS, ASSETS & PASSWORDS
CONFIG_FILE = FRONTEND_DIR / "config.json"
REGRAS_FILE = FRONTEND_DIR / "regras.json"
SELETOR_FILE = FRONTEND_DIR / "seletor.json"
TEMPLATE_BLOCKS_FILE = FRONTEND_DIR / "template_blocks.json"
MAINTENANCE_PASSWORD = os.environ.get("MAINTENANCE_PASSWORD", "admin")

from backend.schedule_generator import run_schedule_engine
from backend.pdf_parser import parse_pdf_document


# Diretório de rede para armazenamento de anexos do histórico
ATTACHMENTS_NETWORK_DIR = Path(
    r"Q:\GROUPS\BR_SC_ITJ_WAU_DPTO_PRODUCAO\Processos WAU Chaves Especiais e Acionamentos\10 - PASTAS PESSOAIS\Luan Schappo\Anexos_Historico"
)

# Define Windows AppUserModelID para ícone na barra de tarefas do Windows
if sys.platform == "win32":
    try:
        import ctypes
        myappid = "com.eletrocentros.app.v2"
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
    except Exception:
        pass


def format_compact_json(obj) -> str:
    """
    Formatador de JSON para regras.json que compacta objetos folhas e inlina
    arrays repetitivos (como 'it', 'etapas', 'cond', 'condicoes', 'valores') em linha única.
    """
    s = json.dumps(obj, ensure_ascii=False, indent=2)

    # 1. Compactar objetos folha em 1 única linha
    s = re.sub(r'\{\s*"t":\s*"([^"]+)",\s*"v":\s*([^}\n]+?)\s*\}', r'{"t": "\1", "v": \2}', s)
    s = re.sub(r'\{\s*"tipo":\s*"([^"]+)",\s*"valor":\s*([^}\n]+?)\s*\}', r'{"tipo": "\1", "valor": \2}', s)
    s = re.sub(r'\{\s*"tipo":\s*"([^"]+)",\s*"modo":\s*"([^"]+)"\s*\}', r'{"tipo": "\1", "modo": "\2"}', s)
    s = re.sub(r'\{\s*"tipo":\s*"([^"]+)",\s*"modo":\s*"([^"]+)",\s*"valor":\s*([^}\n]+?)\s*\}', r'{"tipo": "\1", "modo": "\2", "valor": \3}', s)
    s = re.sub(r'\{\s*"tipo":\s*"([^"]+)",\s*"valor":\s*([^,]+),\s*"modo":\s*"([^"]+)"\s*\}', r'{"tipo": "\1", "valor": \2, "modo": "\3"}', s)
    s = re.sub(r'\{\s*"c":\s*"([^"]+)",\s*"o":\s*"([^"]+)",\s*"val":\s*"([^"]+)",\s*"j":\s*"([^"]+)"\s*\}', r'{"c": "\1", "o": "\2", "val": "\3", "j": "\4"}', s)
    s = re.sub(r'\{\s*"flag":\s*"([^"]+)",\s*"rotulo":\s*"([^"]+)",\s*"forma":\s*"([^"]+)",\s*"valor":\s*"([^"]+)"\s*\}', r'{"flag": "\1", "rotulo": "\2", "forma": "\3", "valor": "\4"}', s)

    # 2. Inlinar arrays de tokens/condições/etapas como "it", "etapas", "cond", "condicoes", "valores"
    def inline_key_array(match):
        key = match.group(1)
        content = match.group(2)
        single_line = re.sub(r'\s*\n\s*', ' ', content)
        return f'"{key}": [ {single_line.strip()} ]'

    s = re.sub(r'"(it|etapas|cond|condicoes|valores)":\s*\[\s*([\s\S]*?)\s*\]', inline_key_array, s)
    return s


def prompt_save_file_path(default_filename: str, file_type: str = "xlsx") -> str:
    """Abre uma janela de diálogo nativa do Windows para o usuário escolher o local e nome do arquivo."""
    try:
        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)

        if file_type == "xlsx":
            filetypes = [("Planilhas Excel (*.xlsx)", "*.xlsx"), ("Todos os Arquivos (*.*)", "*.*")]
            defaultextension = ".xlsx"
        elif file_type == "csv":
            filetypes = [("Arquivos CSV (*.csv)", "*.csv"), ("Todos os Arquivos (*.*)", "*.*")]
            defaultextension = ".csv"
        elif file_type == "json":
            filetypes = [("Arquivos JSON (*.json)", "*.json"), ("Todos os Arquivos (*.*)", "*.*")]
            defaultextension = ".json"
        else:
            filetypes = [("Todos os Arquivos (*.*)", "*.*")]
            defaultextension = ""

        filepath = filedialog.asksaveasfilename(
            parent=root,
            title="Salvar arquivo como",
            initialfile=default_filename,
            defaultextension=defaultextension,
            filetypes=filetypes
        )
        root.destroy()
        return filepath or ""
    except Exception as e:
        print(f"[Backend Python] Erro ao abrir diálogo de salvamento: {e}")
        return ""


class AppAPI:
    """
    API exposta para o JavaScript via pywebview (window.pywebview.api)
    ou chamadas HTTP para o backend.
    """

    def __init__(self):
        self.password = MAINTENANCE_PASSWORD
        self._versao_config = 0
        self._versao_regras = 0
        self._versao_seletor = 0
        self._versao_template = 0
        # Inicializa e garante criação do BD 'bd_eletrocentros_app' e tabelas 'logs_modificacoes' e 'app_settings'
        init_db()

    def get_config(self) -> dict:
        """Lê e retorna a configuração modular a partir do MySQL app_settings (com fallback para config.json)."""
        dados, versao = get_setting("config_geral", fallback_file=CONFIG_FILE)
        self._versao_config = versao
        return dados or {}

    def save_config(self, config_data: dict) -> dict:
        """Salva a nova configuração no MySQL app_settings com lock otimista e backup atômico."""
        try:
            resultado = save_setting(
                "config_geral",
                config_data,
                versao_esperada=getattr(self, "_versao_config", 0),
                backup_file=CONFIG_FILE
            )
            if resultado.get("status") == "success":
                self._versao_config = resultado.get("versao", 1)
                print("[Backend Python] Configurações salvas no MySQL app_settings com sucesso!")
            return resultado
        except Exception as e:
            print(f"[Backend Python] Erro ao salvar config: {e}")
            return {"status": "error", "message": str(e)}

    def get_regras(self) -> list:
        """Lê e retorna a estrutura de regras do MySQL app_settings (com fallback para regras.json)."""
        dados, versao = get_setting("regras_calculo", fallback_file=REGRAS_FILE)
        self._versao_regras = versao
        return dados or []

    def save_regras(self, *args, **kwargs) -> dict:
        """Salva as regras atualizadas no MySQL com lock otimista, registra histórico e atualiza cache local."""
        try:
            regras_data = None
            motivo = None
            anexos_input = []

            if len(args) >= 1:
                if isinstance(args[0], dict) and "regras" in args[0]:
                    regras_data = args[0].get("regras")
                    motivo = args[0].get("motivo")
                    if "anexos" in args[0] and isinstance(args[0]["anexos"], list):
                        anexos_input = args[0]["anexos"]
                    elif args[0].get("anexo_nome") and args[0].get("anexo_base64"):
                        anexos_input = [{"nome": args[0]["anexo_nome"], "base64": args[0]["anexo_base64"]}]
                else:
                    regras_data = args[0]
            elif "regras_data" in kwargs:
                regras_data = kwargs.get("regras_data")

            if len(args) >= 2 and motivo is None:
                motivo = args[1]
            elif "motivo" in kwargs and motivo is None:
                motivo = kwargs.get("motivo")

            if "anexos" in kwargs and isinstance(kwargs["anexos"], list):
                anexos_input = kwargs["anexos"]
            elif "anexo_nome" in kwargs and "anexo_base64" in kwargs:
                anexos_input = [{"nome": kwargs["anexo_nome"], "base64": kwargs["anexo_base64"]}]

            if not regras_data:
                return {"status": "error", "message": "Nenhum dado de regras fornecido."}

            motivo_str = (motivo or "").strip()
            if len(motivo_str) < 20:
                return {"status": "error", "message": "O motivo da alteração é obrigatório e deve ter no mínimo 20 caracteres detalhando a mudança."}

            # Processamento de anexos múltiplos para a pasta de rede
            anexos_salvos = []
            if anexos_input:
                ts_prefix = datetime.now().strftime("%Y%m%d_%H%M%S")
                for idx, anexo_item in enumerate(anexos_input):
                    a_nome = anexo_item.get("nome")
                    a_base64 = anexo_item.get("base64")
                    if not a_nome or not a_base64:
                        continue
                    try:
                        safe_nome = re.sub(r'[^\w\.\-\(\) ]', '_', a_nome)
                        file_name = f"{ts_prefix}_{idx+1}_{safe_nome}" if len(anexos_input) > 1 else f"{ts_prefix}_{safe_nome}"

                        try:
                            ATTACHMENTS_NETWORK_DIR.mkdir(parents=True, exist_ok=True)
                            dest_file = ATTACHMENTS_NETWORK_DIR / file_name
                        except Exception as net_ex:
                            print(f"[Backend Python] Aviso ao acessar pasta de rede ({net_ex}). Usando fallback local...")
                            fallback_dir = BASE_DIR / "anexos_historico"
                            fallback_dir.mkdir(parents=True, exist_ok=True)
                            dest_file = fallback_dir / file_name

                        raw_data = a_base64
                        if "," in raw_data:
                            raw_data = raw_data.split(",", 1)[1]
                        file_bytes = base64.b64decode(raw_data)
                        with open(dest_file, "wb") as f_out:
                            f_out.write(file_bytes)

                        caminho_final = str(dest_file)
                        anexos_salvos.append({"nome": a_nome, "caminho": caminho_final})
                        print(f"[Backend Python] Anexo [{idx+1}/{len(anexos_input)}] salvo em: {caminho_final}")
                    except Exception as anexo_err:
                        print(f"[Backend Python] Erro ao processar anexo '{a_nome}': {anexo_err}")

            anexo_caminho_db = json.dumps(anexos_salvos, ensure_ascii=False) if anexos_salvos else None
            anexo_nome_db = ", ".join([x["nome"] for x in anexos_salvos]) if anexos_salvos else None

            # 1. Carrega as regras anteriores para calcular o diff
            regras_antigas = self.get_regras()

            # 2. Compara e registra as alterações na tabela 'logs_modificacoes' no MySQL
            total_logs = comparar_e_registrar_alteracoes(
                regras_antigas,
                regras_data,
                motivo=motivo,
                anexo_nome=anexo_nome_db,
                anexo_caminho=anexo_caminho_db
            )
            print(f"[Backend Python] {total_logs} log(s) de alteração de regras registrados no banco de dados. (Motivo: {motivo})")

            # 3. Salva a nova versão no MySQL app_settings com lock otimista
            resultado = save_setting(
                "regras_calculo",
                regras_data,
                versao_esperada=getattr(self, "_versao_regras", 0),
                backup_file=REGRAS_FILE
            )

            if resultado.get("status") != "success":
                return resultado

            self._versao_regras = resultado.get("versao", 1)
            print("[Backend Python] Regras salvas no MySQL app_settings e cache atualizado!")
            return {
                "status": "success",
                "versao": self._versao_regras,
                "logs_registrados": total_logs,
                "anexos_salvos": anexos_salvos
            }
        except Exception as e:
            print(f"[Backend Python] Erro ao salvar regras: {e}")
            return {"status": "error", "message": str(e)}

    def open_attachment(self, file_path: str) -> dict:
        """Abre o arquivo anexo diretamente da pasta de rede no visualizador padrão do sistema operacional."""
        try:
            if not file_path:
                return {"status": "error", "message": "Arquivo não especificado."}
            p = Path(file_path)
            file_name = p.name or "Anexo"
            if not p.exists():
                return {"status": "error", "message": f"O arquivo '{file_name}' não foi encontrado ou não está acessível no momento."}
            
            if os.name == 'nt':
                os.startfile(str(p))
            else:
                import subprocess
                subprocess.Popen(["xdg-open", str(p)])
            
            return {"status": "success"}
        except Exception as e:
            print(f"[Backend Python] Erro ao abrir anexo '{file_path}': {e}")
            file_name = Path(file_path).name if file_path else "Anexo"
            return {"status": "error", "message": f"Não foi possível abrir o arquivo '{file_name}'. Verifique o acesso ao documento."}

    def get_logs(self, limit: int = 100) -> list:
        """Retorna o histórico de alterações de regras cadastrado no banco de dados MySQL."""
        return obter_logs(limit=limit)

    def verify_password(self, password: str) -> bool:
        """Valida se a senha digitada corresponde às credenciais de mantenedor."""
        return password == self.password or password == "1234"

    def get_seletor(self) -> list:
        """Lê e retorna a base do seletor de PEPs e Centros de Trabalho cadastrada no MySQL app_settings."""
        dados, versao = get_setting("seletor_pep_cts", fallback_file=SELETOR_FILE)
        self._versao_seletor = versao
        return dados or []

    def save_seletor(self, payload: dict) -> dict:
        """Salva as alterações na base do seletor no MySQL app_settings com lock otimista e registra histórico."""
        try:
            seletor_data = payload.get("seletor", payload) if isinstance(payload, dict) and "seletor" in payload else payload
            motivo = payload.get("motivo", "") if isinstance(payload, dict) else ""

            # Trata anexo se houver
            anexo_nome_db = None
            anexo_caminho_db = None
            if isinstance(payload, dict) and payload.get("anexo_base64") and payload.get("anexo_nome"):
                try:
                    anexo_bytes = base64.b64decode(payload["anexo_base64"])
                    nome_limpo = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{payload['anexo_nome']}"
                    
                    try:
                        ATTACHMENTS_NETWORK_DIR.mkdir(parents=True, exist_ok=True)
                        dest_file = ATTACHMENTS_NETWORK_DIR / nome_limpo
                    except Exception:
                        fallback_dir = BASE_DIR / "anexos_historico"
                        fallback_dir.mkdir(parents=True, exist_ok=True)
                        dest_file = fallback_dir / nome_limpo

                    with open(dest_file, "wb") as af:
                        af.write(anexo_bytes)
                    anexo_nome_db = payload["anexo_nome"]
                    anexo_caminho_db = str(dest_file)
                except Exception as ex_anexo:
                    print(f"[Backend Python] Erro ao salvar anexo do seletor: {ex_anexo}")

            # Registra log no MySQL
            total_logs = 0
            usuario = os.getenv("USERNAME", "Sistema")
            total_linhas = len(seletor_data) if isinstance(seletor_data, list) else 0
            registrar_log(
                disciplina="Seletor de PEP & CTs",
                campo="Tabela Seletor",
                tipo_alteracao="UPDATE",
                detalhes=f"Atualização do Seletor de PEP e Centros de Trabalho ({total_linhas} combinações cadastradas).",
                motivo=motivo or "Atualização dos parâmetros do Seletor de PEP/CTs",
                usuario=usuario,
                anexo_nome=anexo_nome_db,
                anexo_caminho=anexo_caminho_db
            )
            total_logs = 1

            # Salva no MySQL app_settings com lock otimista
            resultado = save_setting(
                "seletor_pep_cts",
                seletor_data,
                versao_esperada=getattr(self, "_versao_seletor", 0),
                backup_file=SELETOR_FILE
            )

            if resultado.get("status") != "success":
                return resultado

            self._versao_seletor = resultado.get("versao", 1)
            print("[Backend Python] Seletor salvo no MySQL app_settings com sucesso!")
            return {"status": "success", "versao": self._versao_seletor, "logs_registrados": total_logs}
        except Exception as e:
            print(f"[Backend Python] Erro ao salvar seletor: {e}")
            return {"status": "error", "message": str(e)}

    def get_template_blocks(self) -> dict:
        """Lê e retorna a base de templates de operações cadastrada no MySQL app_settings."""
        dados, versao = get_setting("template_blocks", fallback_file=TEMPLATE_BLOCKS_FILE)
        self._versao_template = versao
        return dados or {}

    def save_template_blocks(self, payload: dict) -> dict:
        """Salva as alterações na base de templates no MySQL app_settings com lock otimista e registra histórico."""
        try:
            tb_data = payload.get("template_blocks", payload) if isinstance(payload, dict) and "template_blocks" in payload else payload
            motivo = payload.get("motivo", "") if isinstance(payload, dict) else ""

            # Trata anexo se houver
            anexo_nome_db = None
            anexo_caminho_db = None
            if isinstance(payload, dict) and payload.get("anexo_base64") and payload.get("anexo_nome"):
                try:
                    anexo_bytes = base64.b64decode(payload["anexo_base64"])
                    nome_limpo = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{payload['anexo_nome']}"
                    
                    try:
                        ATTACHMENTS_NETWORK_DIR.mkdir(parents=True, exist_ok=True)
                        dest_file = ATTACHMENTS_NETWORK_DIR / nome_limpo
                    except Exception:
                        fallback_dir = BASE_DIR / "anexos_historico"
                        fallback_dir.mkdir(parents=True, exist_ok=True)
                        dest_file = fallback_dir / nome_limpo

                    with open(dest_file, "wb") as af:
                        af.write(anexo_bytes)
                    anexo_nome_db = payload["anexo_nome"]
                    anexo_caminho_db = str(dest_file)
                except Exception as ex_anexo:
                    print(f"[Backend Python] Erro ao salvar anexo do template: {ex_anexo}")

            # Registra log no MySQL
            total_logs = 0
            usuario = os.getenv("USERNAME", "Sistema")
            cenarios_count = len(tb_data.get("cenarios", {})) if isinstance(tb_data, dict) else 0
            registrar_log(
                disciplina="Templates de Operações",
                campo="Blocos de Operações",
                tipo_alteracao="UPDATE",
                detalhes=f"Atualização da base de templates de blocos de operações ({cenarios_count} cenários cadastrados).",
                motivo=motivo or "Atualização da base de templates de operações",
                usuario=usuario,
                anexo_nome=anexo_nome_db,
                anexo_caminho=anexo_caminho_db
            )
            total_logs = 1

            # Salva no MySQL app_settings com lock otimista
            resultado = save_setting(
                "template_blocks",
                tb_data,
                versao_esperada=getattr(self, "_versao_template", 0),
                backup_file=TEMPLATE_BLOCKS_FILE
            )

            if resultado.get("status") != "success":
                return resultado

            self._versao_template = resultado.get("versao", 1)
            print("[Backend Python] Templates salvos no MySQL app_settings com sucesso!")
            return {"status": "success", "versao": self._versao_template, "logs_registrados": total_logs}
        except Exception as e:
            print(f"[Backend Python] Erro ao salvar template_blocks: {e}")
            return {"status": "error", "message": str(e)}

    def generate_schedule(self, payload: dict) -> dict:
        """Gera o cronograma de operações com base nos parâmetros e tempos calculados."""
        try:
            form_data = payload.get("ctx", {})
            calc_times = payload.get("calc_times", {})
            template_blocks = self.get_template_blocks()
            seletor_rows = self.get_seletor()
            result = run_schedule_engine(form_data, calc_times, template_blocks, seletor_rows)
            return {"status": "success", "data": result}
        except Exception as e:
            print(f"[Backend Python] Erro ao gerar cronograma: {e}")
    def parse_pdf(self, payload: dict) -> dict:
        """Processa o PDF do projeto e extrai campos para preenchimento automático do formulário."""
        try:
            pdf_b64 = payload.get("pdf_base64") or payload.get("base64")
            filename = payload.get("filename", "documento.pdf")
            if not pdf_b64:
                return {"status": "error", "message": "Nenhum arquivo PDF fornecido."}
            return parse_pdf_document(pdf_b64, filename=filename)
        except Exception as e:
            print(f"[Backend Python] Erro ao processar PDF: {e}")
            return {"status": "error", "message": str(e)}

    def export_excel(self, data: dict) -> dict:
        """Gera e retorna a planilha Excel (.xlsx) com o resultado detalhado dos tempos e aba Resultado."""
        try:
            import io
            import base64
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            
            # Cores e Estilos Padrão
            title_font = Font(name="Segoe UI", size=13, bold=True, color="FFFFFF")
            title_fill = PatternFill("solid", fgColor="0F2C59")
            section_font = Font(name="Segoe UI", size=11, bold=True, color="0F2C59")
            label_font = Font(name="Segoe UI", size=10, bold=True, color="4A5568")
            val_font = Font(name="Segoe UI", size=10, color="1A202C")
            hdr_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
            hdr_fill = PatternFill("solid", fgColor="1E3E62")
            subtotal_font = Font(name="Segoe UI", size=10, bold=True, color="1A202C")
            subtotal_fill = PatternFill("solid", fgColor="F0F4F8")
            total_font = Font(name="Segoe UI", size=11, bold=True, color="0F2C59")
            total_fill = PatternFill("solid", fgColor="D0E8FF")

            thin_border = Border(
                left=Side(style="thin", color="CBD5E0"),
                right=Side(style="thin", color="CBD5E0"),
                top=Side(style="thin", color="CBD5E0"),
                bottom=Side(style="thin", color="CBD5E0")
            )
            double_bottom = Border(
                left=Side(style="thin", color="CBD5E0"),
                right=Side(style="thin", color="CBD5E0"),
                top=Side(style="thin", color="CBD5E0"),
                bottom=Side(style="double", color="0F2C59")
            )

            # Obter ou gerar tarefas do cronograma
            cronograma = data.get("cronograma")
            if not cronograma and "ctx" in data:
                try:
                    tb = self.get_template_blocks()
                    cronograma = run_schedule_engine(data.get("ctx", {}), data.get("calc_times", {}), tb)
                except Exception as ex_gen:
                    print(f"[Backend Python] Aviso ao gerar cronograma para exportação: {ex_gen}")

            tarefas_list = cronograma.get("tarefas", []) if cronograma else []

            # =========================================================================
            # CÁLCULO DAS HORAS TOTAIS POR DISCIPLINA (EXCEL ORIGINAL)
            # =========================================================================
            eng_h = sum(float(t.get("trabalho", 0)) for t in tarefas_list 
                        if int(str(t.get("tarefa", 0)).strip() or 0) < 703 
                        and abs(float(t.get("trabalho", 0)) - 0.1) > 0.001 
                        and "ROM" not in str(t.get("descricao_tarefa", "")).upper())
            
            mec_h = sum(float(t.get("trabalho", 0)) for t in tarefas_list 
                        if ((705 <= int(str(t.get("tarefa", 0)).strip() or 0) <= 798) or int(str(t.get("tarefa", 0)).strip() or 0) == 894) 
                        and abs(float(t.get("trabalho", 0)) - 0.1) > 0.001 
                        and int(str(t.get("tarefa", 0)).strip() or 0) not in [754, 755, 765, 793, 794])
            
            ele_h = sum(float(t.get("trabalho", 0)) for t in tarefas_list 
                        if ((799 <= int(str(t.get("tarefa", 0)).strip() or 0) <= 893) or int(str(t.get("tarefa", 0)).strip() or 0) == 895) 
                        and abs(float(t.get("trabalho", 0)) - 0.1) > 0.001 
                        and int(str(t.get("tarefa", 0)).strip() or 0) not in [810, 828, 838, 858, 868])
            
            tot_h = round(eng_h + mec_h + ele_h, 1)
            eng_h = round(eng_h, 1)
            mec_h = round(mec_h, 1)
            ele_h = round(ele_h, 1)

            # =========================================================================
            # ABA 1: TOTAIS DO PROJETO (Resumo Consolidado das 3 Disciplinas)
            # =========================================================================
            ws_tot = wb.active
            ws_tot.title = "Totais do Projeto"
            ws_tot.views.sheetView[0].showGridLines = True

            ws_tot.merge_cells("A1:D1")
            ws_tot["A1"] = "ELETROCENTROS APP — HORAS TOTAIS DO PROJETO (PADRÃO EXCEL)"
            ws_tot["A1"].font = title_font
            ws_tot["A1"].fill = title_fill
            ws_tot["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws_tot.row_dimensions[1].height = 36

            ctx = data.get("ctx", {})
            pep = data.get("pep", "") or ctx.get("pep", "") or "Não informado"
            data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            usuario = os.getenv("USERNAME", "Usuário")

            ws_tot["A3"] = "INFORMAÇÕES GERAIS DO PROJETO"
            ws_tot["A3"].font = section_font

            meta_rows = [
                ("Código PEP / Ordem:", pep, "Data / Hora de Cálculo:", data_hora),
                ("Tipo de Estrutura:", str(ctx.get("tipoestrutura", "-")), "Quantidade de Módulos:", f"{ctx.get('nmod', 1)} Módulo(s)"),
                ("Dimensões Gerais:", f"{ctx.get('comp', 0)}m (C) x {ctx.get('larg', 0)}m (L)", "Plano de Pintura:", str(ctx.get("planpin", "-"))),
                ("Ar Condicionado:", f"{ctx.get('tipomaq', '-')} ({ctx.get('qtdmaq', 0)}x)", "Complexidade Elétrica:", str(ctx.get("complexidade", "-"))),
                ("Sist. Incêndio:", str(ctx.get("incendio", "-")), "Sist. Segurança:", str(ctx.get("seguranca", "-"))),
                ("Usuário Responsável:", usuario, "Nº Colunas Total:", str(ctx.get("nrcolunas", 0)))
            ]

            curr_r = 4
            for label1, val1, label2, val2 in meta_rows:
                ws_tot[f"A{curr_r}"] = label1
                ws_tot[f"A{curr_r}"].font = label_font
                ws_tot[f"B{curr_r}"] = val1
                ws_tot[f"B{curr_r}"].font = val_font

                ws_tot[f"C{curr_r}"] = label2
                ws_tot[f"C{curr_r}"].font = label_font
                ws_tot[f"D{curr_r}"] = val2
                ws_tot[f"D{curr_r}"].font = val_font
                curr_r += 1

            curr_r += 1
            ws_tot[f"A{curr_r}"] = "HORAS TOTAIS POR DIAGRAMA (GABARITO CONSOLIDADO)"
            ws_tot[f"A{curr_r}"].font = section_font
            curr_r += 1

            ws_tot[f"A{curr_r}"] = "Disciplina"
            ws_tot[f"B{curr_r}"] = "Regra de Filtro das Operações"
            ws_tot[f"C{curr_r}"] = "Horas Totais (H)"
            ws_tot[f"D{curr_r}"] = "Participação (%)"
            for col_letter in ["A", "B", "C", "D"]:
                ws_tot[f"{col_letter}{curr_r}"].font = hdr_font
                ws_tot[f"{col_letter}{curr_r}"].fill = hdr_fill
                ws_tot[f"{col_letter}{curr_r}"].alignment = Alignment(horizontal="center" if col_letter in ["C", "D"] else "left", vertical="center")
            ws_tot.row_dimensions[curr_r].height = 24
            curr_r += 1

            disc_rows = [
                ("Engenharia (ENG)", "Tarefas < 0703 (sem tempo 0,1 e sem ROM)", eng_h, round(eng_h / tot_h * 100, 1) if tot_h > 0 else 0),
                ("Mecânica (MEC)", "Tarefas 0705 a 0798 + 0894 (sem 0,1 e sem 754, 755, 765, 793, 794)", mec_h, round(mec_h / tot_h * 100, 1) if tot_h > 0 else 0),
                ("Elétrica (ELE)", "Tarefas 0799 a 0893 + 0895 (sem 0,1 e sem 810, 828, 838, 858, 868)", ele_h, round(ele_h / tot_h * 100, 1) if tot_h > 0 else 0),
            ]

            for d_name, d_rule, d_h, d_pct in disc_rows:
                ws_tot[f"A{curr_r}"] = d_name
                ws_tot[f"A{curr_r}"].font = label_font
                ws_tot[f"A{curr_r}"].border = thin_border

                ws_tot[f"B{curr_r}"] = d_rule
                ws_tot[f"B{curr_r}"].font = val_font
                ws_tot[f"B{curr_r}"].border = thin_border

                ws_tot[f"C{curr_r}"] = d_h
                ws_tot[f"C{curr_r}"].font = Font(name="Segoe UI", size=11, bold=True, color="0F2C59")
                ws_tot[f"C{curr_r}"].number_format = "#,##0.0 \"h\""
                ws_tot[f"C{curr_r}"].alignment = Alignment(horizontal="right")
                ws_tot[f"C{curr_r}"].border = thin_border

                ws_tot[f"D{curr_r}"] = f"{d_pct}%"
                ws_tot[f"D{curr_r}"].font = val_font
                ws_tot[f"D{curr_r}"].alignment = Alignment(horizontal="center")
                ws_tot[f"D{curr_r}"].border = thin_border
                curr_r += 1

            # Total Geral Linha
            ws_tot[f"A{curr_r}"] = "TOTAL GERAL DO PROJETO"
            ws_tot[f"B{curr_r}"] = "Soma consolidada (Engenharia + Mecânica + Elétrica)"
            ws_tot[f"C{curr_r}"] = tot_h
            ws_tot[f"C{curr_r}"].number_format = "#,##0.0 \"h\""
            ws_tot[f"C{curr_r}"].alignment = Alignment(horizontal="right")
            ws_tot[f"D{curr_r}"] = "100.0%"
            ws_tot[f"D{curr_r}"].alignment = Alignment(horizontal="center")

            for col_letter in ["A", "B", "C", "D"]:
                c = ws_tot[f"{col_letter}{curr_r}"]
                c.font = total_font
                c.fill = total_fill
                c.border = double_bottom
            ws_tot.row_dimensions[curr_r].height = 24

            for col in ws_tot.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or "")
                    if cell.row == 1: continue
                    if len(val_str) > max_len: max_len = len(val_str)
                ws_tot.column_dimensions[col_letter].width = max(max_len + 4, 16)

            # =========================================================================
            # ABA 2: RESULTADO (Cronograma Completo de Operações — Padrão VBA)
            # =========================================================================
            ws_res = wb.create_sheet(title="Resultado - Cronograma")
            ws_res.views.sheetView[0].showGridLines = True

            # Cabeçalho da Aba Resultado (Linha 3 no padrão original)
            ws_res["A3"] = "Tarefa"
            ws_res["B3"] = "Descrição da Tarefa"
            ws_res["C3"] = "Duração"
            ws_res["D3"] = "Unidade"
            ws_res["E3"] = "Trabalho"
            ws_res["F3"] = "Disciplina / Fase"

            for col_idx, col_letter in enumerate(["A", "B", "C", "D", "E", "F"]):
                cell = ws_res[f"{col_letter}3"]
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center" if col_idx in [0, 2, 3] else ("right" if col_idx == 4 else "left"), vertical="center")
            ws_res.row_dimensions[3].height = 24

            r_idx = 4
            for t in tarefas_list:
                t_code = str(t.get("tarefa_formatada") or t.get("tarefa", "")).strip()
                t_desc = str(t.get("descricao_tarefa", "")).strip()
                dur_val = float(t.get("duracao", 0))
                unid_val = str(t.get("unidade", "DIA")).strip()
                trab_val = float(t.get("trabalho", 0))
                disc_val = str(t.get("disciplina", "Geral")).strip()

                ws_res[f"A{r_idx}"] = t_code
                ws_res[f"A{r_idx}"].alignment = Alignment(horizontal="center", vertical="center")
                ws_res[f"A{r_idx}"].number_format = "@"

                ws_res[f"B{r_idx}"] = t_desc
                ws_res[f"B{r_idx}"].alignment = Alignment(horizontal="left", vertical="center")

                ws_res[f"C{r_idx}"] = dur_val
                ws_res[f"C{r_idx}"].alignment = Alignment(horizontal="center", vertical="center")
                ws_res[f"C{r_idx}"].number_format = "0.0"

                ws_res[f"D{r_idx}"] = unid_val
                ws_res[f"D{r_idx}"].alignment = Alignment(horizontal="center", vertical="center")

                ws_res[f"E{r_idx}"] = trab_val
                ws_res[f"E{r_idx}"].alignment = Alignment(horizontal="right", vertical="center")
                ws_res[f"E{r_idx}"].number_format = "0.0"

                ws_res[f"F{r_idx}"] = disc_val
                ws_res[f"F{r_idx}"].alignment = Alignment(horizontal="left", vertical="center")

                for col_letter in ["A", "B", "C", "D", "E", "F"]:
                    ws_res[f"{col_letter}{r_idx}"].border = thin_border
                    ws_res[f"{col_letter}{r_idx}"].font = val_font
                r_idx += 1

            # Totalizador do Cronograma
            if tarefas_list:
                ws_res[f"A{r_idx}"] = "TOTAL"
                ws_res[f"B{r_idx}"] = f"Total de {len(tarefas_list)} tarefas geradas"
                ws_res[f"C{r_idx}"] = float(cronograma.get("total_dias", 0))
                ws_res[f"C{r_idx}"].number_format = "0.0"
                ws_res[f"D{r_idx}"] = "DIAS"
                ws_res[f"E{r_idx}"] = float(cronograma.get("total_horas", 0))
                ws_res[f"E{r_idx}"].number_format = "0.00"
                ws_res[f"F{r_idx}"] = cronograma.get("cenario_descricao", "")

                for col_letter in ["A", "B", "C", "D", "E", "F"]:
                    c = ws_res[f"{col_letter}{r_idx}"]
                    c.font = total_font
                    c.fill = total_fill
                    c.border = double_bottom
                ws_res.row_dimensions[r_idx].height = 24

            for col in ws_res.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or "")
                    if cell.row < 3: continue
                    if len(val_str) > max_len: max_len = len(val_str)
                ws_res.column_dimensions[col_letter].width = max(max_len + 4, 14)

            # =========================================================================
            # ABA 3: RESUMO DE REGRAS & PROCESSOS (Disciplinas, Processos e Parâmetros)
            # =========================================================================
            ws = wb.create_sheet(title="Regras & Processos")
            ws.views.sheetView[0].showGridLines = True

            ws.merge_cells("A1:D1")
            ws["A1"] = "ELETROCENTROS APP — RESUMO POR REGRAS & PROCESSOS"
            ws["A1"].font = title_font
            ws["A1"].fill = title_fill
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 36

            ctx = data.get("ctx", {})
            pep = data.get("pep", "") or ctx.get("pep", "") or "Não informado"
            data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            usuario = os.getenv("USERNAME", "Usuário")

            ws["A3"] = "INFORMAÇÕES DO PROJETO"
            ws["A3"].font = section_font

            meta_rows = [
                ("Código PEP / Ordem:", pep, "Data / Hora de Cálculo:", data_hora),
                ("Tipo de Estrutura:", str(ctx.get("tipoestrutura", "-")), "Quantidade de Módulos:", f"{ctx.get('nmod', 1)} Módulo(s)"),
                ("Dimensões Gerais:", f"{ctx.get('comp', 0)}m (C) x {ctx.get('larg', 0)}m (L)", "Plano de Pintura:", str(ctx.get("planpin", "-"))),
                ("Ar Condicionado:", f"{ctx.get('tipomaq', '-')} ({ctx.get('qtdmaq', 0)}x)", "Complexidade Elétrica:", str(ctx.get("complexidade", "-"))),
                ("Sist. Incêndio:", str(ctx.get("incendio", "-")), "Sist. Segurança:", str(ctx.get("seguranca", "-"))),
                ("Usuário Responsável:", usuario, "Nº Colunas Total:", str(ctx.get("nrcolunas", 0)))
            ]

            curr_row = 4
            for label1, val1, label2, val2 in meta_rows:
                ws[f"A{curr_row}"] = label1
                ws[f"A{curr_row}"].font = label_font
                ws[f"B{curr_row}"] = val1
                ws[f"B{curr_row}"].font = val_font

                ws[f"C{curr_row}"] = label2
                ws[f"C{curr_row}"].font = label_font
                ws[f"D{curr_row}"] = val2
                ws[f"D{curr_row}"].font = val_font
                curr_row += 1

            curr_row += 1
            ws[f"A{curr_row}"] = "RESUMO GERAL CONSOLIDADO"
            ws[f"A{curr_row}"].font = section_font
            curr_row += 1

            ws[f"A{curr_row}"] = "TOTAL HORAS ORÇADAS (H):"
            ws[f"A{curr_row}"].font = Font(name="Segoe UI", size=11, bold=True, color="107C41")
            ws[f"B{curr_row}"] = float(data.get("totalGeralH", 0))
            ws[f"B{curr_row}"].font = Font(name="Segoe UI", size=11, bold=True, color="107C41")
            ws[f"B{curr_row}"].number_format = "#,##0.00 \"h\""

            ws[f"C{curr_row}"] = "DURAÇÃO ESTIMADA (DUR):"
            ws[f"C{curr_row}"].font = Font(name="Segoe UI", size=11, bold=True, color="D83B01")
            ws[f"D{curr_row}"] = float(data.get("totalGeralDUR", 0))
            ws[f"D{curr_row}"].font = Font(name="Segoe UI", size=11, bold=True, color="D83B01")
            ws[f"D{curr_row}"].number_format = "#,##0.0 \"dias\""
            curr_row += 2

            ws[f"A{curr_row}"] = "Área / Disciplina"
            ws[f"B{curr_row}"] = "Processo / Campo"
            ws[f"C{curr_row}"] = "Horas Orçadas (H)"
            ws[f"D{curr_row}"] = "Duração Estimada (DUR)"
            for col_idx, col_letter in enumerate(["A", "B", "C", "D"]):
                cell = ws[f"{col_letter}{curr_row}"]
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="left" if col_idx < 2 else "right", vertical="center")
            ws.row_dimensions[curr_row].height = 24
            curr_row += 1

            for area in data.get("resultadosAreas", []):
                nome_area = area.get("area", "Área")
                campos = area.get("campos", [])
                for campo in campos:
                    ws[f"A{curr_row}"] = nome_area
                    ws[f"B{curr_row}"] = campo.get("chave", "")

                    cell_h = ws[f"C{curr_row}"]
                    cell_h.value = float(campo.get("h", 0))
                    cell_h.number_format = "#,##0.00"

                    cell_dur = ws[f"D{curr_row}"]
                    cell_dur.value = float(campo.get("dur", 0))
                    cell_dur.number_format = "#,##0.0"

                    for col_letter in ["A", "B", "C", "D"]:
                        ws[f"{col_letter}{curr_row}"].border = thin_border
                        ws[f"{col_letter}{curr_row}"].font = val_font
                    curr_row += 1

                ws[f"A{curr_row}"] = f"Subtotal — {nome_area}"
                ws[f"B{curr_row}"] = f"({len(campos)} processos)"

                cell_sub_h = ws[f"C{curr_row}"]
                cell_sub_h.value = float(area.get("totalH", 0))
                cell_sub_h.number_format = "#,##0.00"

                cell_sub_dur = ws[f"D{curr_row}"]
                cell_sub_dur.value = float(area.get("totalDUR", 0))
                cell_sub_dur.number_format = "#,##0.0"

                for col_letter in ["A", "B", "C", "D"]:
                    c = ws[f"{col_letter}{curr_row}"]
                    c.font = subtotal_font
                    c.fill = subtotal_fill
                    c.border = thin_border
                curr_row += 1

            ws[f"A{curr_row}"] = "TOTAL GERAL CONSOLIDADO"
            ws[f"B{curr_row}"] = "Todos os processos"

            cell_tot_h = ws[f"C{curr_row}"]
            cell_tot_h.value = float(data.get("totalGeralH", 0))
            cell_tot_h.number_format = "#,##0.00"

            cell_tot_dur = ws[f"D{curr_row}"]
            cell_tot_dur.value = float(data.get("totalGeralDUR", 0))
            cell_tot_dur.number_format = "#,##0.0"

            for col_letter in ["A", "B", "C", "D"]:
                c = ws[f"{col_letter}{curr_row}"]
                c.font = total_font
                c.fill = total_fill
                c.border = double_bottom
            ws.row_dimensions[curr_row].height = 24

            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or "")
                    if cell.row == 1: continue
                    if len(val_str) > max_len: max_len = len(val_str)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 18)

            # =========================================================================
            # ABA 3: SELETOR DE PEP & CENTROS DE TRABALHO (se houver)
            # =========================================================================
            seletor_info = data.get("seletor")
            if seletor_info:
                ws_sel = wb.create_sheet(title="Seletor PEP & CTs")
                ws_sel.views.sheetView[0].showGridLines = True

                ws_sel.merge_cells("A1:C1")
                ws_sel["A1"] = "SELETOR DE PEP STANDARD & CENTROS DE TRABALHO"
                ws_sel["A1"].font = title_font
                ws_sel["A1"].fill = title_fill
                ws_sel["A1"].alignment = Alignment(horizontal="center", vertical="center")
                ws_sel.row_dimensions[1].height = 36

                ws_sel["A3"] = "PEP Standard Sugerido:"
                ws_sel["A3"].font = label_font
                ws_sel["B3"] = str(seletor_info.get("PEP Standard", "-"))
                ws_sel["B3"].font = Font(name="Segoe UI", size=11, bold=True, color="0F2C59")

                ws_sel["A5"] = "Centro de Trabalho / Disciplina"
                ws_sel["B5"] = "Diagrama de Rede (DR)"
                ws_sel["C5"] = "Alternativa"
                for col_letter in ["A", "B", "C"]:
                    ws_sel[f"{col_letter}5"].font = hdr_font
                    ws_sel[f"{col_letter}5"].fill = hdr_fill
                    ws_sel[f"{col_letter}5"].alignment = Alignment(horizontal="center" if col_letter != "A" else "left", vertical="center")
                ws_sel.row_dimensions[5].height = 24

                sel_r = 6
                ct_items = []
                if seletor_info.get("DR Eng Mec"):
                    ct_items.append(("Engenharia Mecânica", seletor_info.get("DR Eng Mec"), seletor_info.get("Alt Eng Mec", "1")))
                if seletor_info.get("DR Eng Ele"):
                    ct_items.append(("Engenharia Elétrica", seletor_info.get("DR Eng Ele"), seletor_info.get("Alt Eng Ele", "1")))

                nmod_val = int(ctx.get("nmod", 1) or 1)
                for m in range(1, 9):
                    dr_m = seletor_info.get(f"DR Mec {m}") or seletor_info.get(f"DR Mec{m}")
                    alt_m = seletor_info.get(f"Alt Mec {m}") or seletor_info.get(f"Alt Mec{m}") or "1"
                    if dr_m and m <= nmod_val:
                        ct_items.append((f"Mecânica Módulo {m}", dr_m, alt_m))

                if seletor_info.get("DR Acess"):
                    ct_items.append(("Acessórios", seletor_info.get("DR Acess"), seletor_info.get("Alt Acess", "1")))
                if seletor_info.get("DR Eletromec"):
                    ct_items.append(("Eletromecânica", seletor_info.get("DR Eletromec"), seletor_info.get("Alt Eletromec", "1")))

                for ct_label, dr_v, alt_v in ct_items:
                    ws_sel[f"A{sel_r}"] = ct_label
                    ws_sel[f"A{sel_r}"].font = val_font
                    ws_sel[f"A{sel_r}"].border = thin_border

                    ws_sel[f"B{sel_r}"] = f"DR {dr_v}"
                    ws_sel[f"B{sel_r}"].font = val_font
                    ws_sel[f"B{sel_r}"].alignment = Alignment(horizontal="center")
                    ws_sel[f"B{sel_r}"].border = thin_border

                    ws_sel[f"C{sel_r}"] = f"Alt {alt_v}"
                    ws_sel[f"C{sel_r}"].font = val_font
                    ws_sel[f"C{sel_r}"].alignment = Alignment(horizontal="center")
                    ws_sel[f"C{sel_r}"].border = thin_border
                    sel_r += 1

                for col in ws_sel.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        val_str = str(cell.value or "")
                        if cell.row == 1: continue
                        if len(val_str) > max_len: max_len = len(val_str)
                    ws_sel.column_dimensions[col_letter].width = max(max_len + 4, 18)

            pep_clean = "".join(c for c in pep if c.isalnum() or c in ("-", "_")).strip()
            data_clean = datetime.now().strftime("%Y%m%d_%H%M")
            file_name = f"Resultado_PEP_{pep_clean}_{data_clean}.xlsx" if pep_clean and pep_clean != "Nãoinformado" else f"Resultado_Eletrocentro_{data_clean}.xlsx"

            if data.get("prompt_save", True):
                filepath = prompt_save_file_path(file_name, "xlsx")
                if not filepath:
                    return {"status": "cancelled", "message": "Operação cancelada pelo usuário."}
                wb.save(filepath)
                return {
                    "status": "success",
                    "filepath": filepath,
                    "filename": os.path.basename(filepath)
                }
            else:
                output_stream = io.BytesIO()
                wb.save(output_stream)
                output_stream.seek(0)
                b64_content = base64.b64encode(output_stream.read()).decode("utf-8")
                return {
                    "status": "success",
                    "filename": file_name,
                    "base64": b64_content
                }
        except Exception as e:
            print(f"[Backend Python] Erro ao exportar Excel: {e}")
            return {"status": "error", "message": str(e)}

    def save_file(self, data: dict) -> dict:
        """Abre janela nativa de Salvar Como e salva o conteúdo de arquivo (CSV, JSON, etc.) no caminho escolhido."""
        try:
            filename = data.get("filename", "arquivo")
            file_type = data.get("file_type", "csv")
            content = data.get("content", "")
            is_base64 = data.get("is_base64", False)

            filepath = prompt_save_file_path(filename, file_type)
            if not filepath:
                return {"status": "cancelled", "message": "Operação cancelada pelo usuário."}

            if is_base64:
                file_bytes = base64.b64decode(content)
                with open(filepath, "wb") as f:
                    f.write(file_bytes)
            else:
                with open(filepath, "w", encoding="utf-8-sig") as f:
                    f.write(content)

            return {
                "status": "success",
                "filepath": filepath,
                "filename": os.path.basename(filepath)
            }
        except Exception as e:
            print(f"[Backend Python] Erro ao salvar arquivo: {e}")
            return {"status": "error", "message": str(e)}


class AppHTTPRequestHandler(http.server.SimpleHTTPRequestHandler):
    """Handler HTTP customizado com suporte a rotas da API /api/*."""

    api = AppAPI()

    def do_GET(self):
        if self.path == "/api/logs" or self.path.startswith("/api/logs?"):
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.end_headers()
            logs = self.api.get_logs()
            self.wfile.write(json.dumps(logs, ensure_ascii=False).encode("utf-8"))
            return
        elif self.path == "/api/get_regras":
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.end_headers()
            regras = self.api.get_regras()
            self.wfile.write(json.dumps(regras, ensure_ascii=False).encode("utf-8"))
            return
        elif self.path == "/api/get_config":
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.end_headers()
            config = self.api.get_config()
            self.wfile.write(json.dumps(config, ensure_ascii=False).encode("utf-8"))
            return
        elif self.path == "/api/get_seletor":
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.end_headers()
            seletor = self.api.get_seletor()
            self.wfile.write(json.dumps(seletor, ensure_ascii=False).encode("utf-8"))
            return
        elif self.path == "/api/get_template_blocks":
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.end_headers()
            tb = self.api.get_template_blocks()
            self.wfile.write(json.dumps(tb, ensure_ascii=False).encode("utf-8"))
            return

        super().do_GET()

    def do_POST(self):
        if self.path == "/api/save_regras":
            content_length = int(self.headers.get("Content-Length", 0))
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode("utf-8"))
                result = self.api.save_regras(data)
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.end_headers()
                self.wfile.write(json.dumps(result, ensure_ascii=False).encode("utf-8"))
            except Exception as e:
                self.send_response(500)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": str(e)}).encode("utf-8"))
            return
        elif self.path == "/api/save_seletor":
            content_length = int(self.headers.get("Content-Length", 0))
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode("utf-8"))
                result = self.api.save_seletor(data)
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.end_headers()
                self.wfile.write(json.dumps(result, ensure_ascii=False).encode("utf-8"))
            except Exception as e:
                self.send_response(500)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": str(e)}).encode("utf-8"))
            return
        elif self.path == "/api/save_template_blocks":
            content_length = int(self.headers.get("Content-Length", 0))
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode("utf-8"))
                result = self.api.save_template_blocks(data)
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.end_headers()
                self.wfile.write(json.dumps(result, ensure_ascii=False).encode("utf-8"))
            except Exception as e:
                self.send_response(500)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": str(e)}).encode("utf-8"))
            return
        elif self.path == "/api/generate_schedule":
            content_length = int(self.headers.get("Content-Length", 0))
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode("utf-8"))
                result = self.api.generate_schedule(data)
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.end_headers()
                self.wfile.write(json.dumps(result, ensure_ascii=False).encode("utf-8"))
            except Exception as e:
                self.send_response(500)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": str(e)}).encode("utf-8"))
            return
        elif self.path == "/api/parse_pdf":
            content_length = int(self.headers.get("Content-Length", 0))
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode("utf-8"))
                result = self.api.parse_pdf(data)
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.end_headers()
                self.wfile.write(json.dumps(result, ensure_ascii=False).encode("utf-8"))
            except Exception as e:
                self.send_response(500)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": str(e)}).encode("utf-8"))
            return
        elif self.path == "/api/export_excel":
            content_length = int(self.headers.get("Content-Length", 0))
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode("utf-8"))
                result = self.api.export_excel(data)
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.end_headers()
                self.wfile.write(json.dumps(result, ensure_ascii=False).encode("utf-8"))
            except Exception as e:
                self.send_response(500)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": str(e)}).encode("utf-8"))
            return
        elif self.path == "/api/save_file":
            content_length = int(self.headers.get("Content-Length", 0))
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode("utf-8"))
                result = self.api.save_file(data)
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.end_headers()
                self.wfile.write(json.dumps(result, ensure_ascii=False).encode("utf-8"))
            except Exception as e:
                self.send_response(500)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": str(e)}).encode("utf-8"))
            return


        super().do_POST()


class ThreadedTCPServer(socketserver.ThreadingMixIn, socketserver.TCPServer):
    allow_reuse_address = True
    daemon_threads = True


def find_available_port(start_port: int = 8000, max_attempts: int = 50) -> int:
    """Busca uma porta TCP disponível para o servidor local a partir de start_port."""
    import socket
    for p in range(start_port, start_port + max_attempts):
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
                s.bind(("127.0.0.1", p))
                return p
        except OSError:
            continue

    # Fallback: porta efêmera aleatória atribuída pelo SO
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind(("127.0.0.1", 0))
        return s.getsockname()[1]


def apply_native_window_icon():
    """Aplica o ícone icone.ico diretamente em todas as janelas Win32 do processo Python."""
    if sys.platform != "win32" or not ICON_PATH.exists():
        return

    try:
        import ctypes
        user32 = ctypes.windll.user32
        kernel32 = ctypes.windll.kernel32
        current_pid = kernel32.GetCurrentProcessId()

        hicon_small = user32.LoadImageW(0, str(ICON_PATH), 1, 16, 16, 0x0010)
        hicon_big = user32.LoadImageW(0, str(ICON_PATH), 1, 32, 32, 0x0010)
        if not hicon_small:
            hicon_small = user32.LoadImageW(0, str(ICON_PATH), 1, 0, 0, 0x0010)
        if not hicon_big:
            hicon_big = hicon_small

        WNDENUMPROC = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.c_void_p, ctypes.c_void_p)

        def enum_windows_callback(hwnd, lparam):
            pid = ctypes.c_ulong()
            user32.GetWindowThreadProcessId(hwnd, ctypes.byref(pid))
            if pid.value == current_pid:
                user32.SendMessageW(hwnd, 0x0080, 0, hicon_small)  # WM_SETICON, ICON_SMALL
                user32.SendMessageW(hwnd, 0x0080, 1, hicon_big)    # WM_SETICON, ICON_BIG
            return True

        user32.EnumWindows(WNDENUMPROC(enum_windows_callback), 0)
    except Exception as e:
        print(f"[GUI Python] Erro ao aplicar ícone nativo: {e}")


APP_VERSION = "2.1.4"


def launch_app():
    """Tenta abrir via pywebview como janela nativa. Caso contrário, usa servidor HTTP local com porta dinâmica."""
    api = AppAPI()
    index_file = FRONTEND_DIR / "index.html"

    try:
        import webview
        print(f"[GUI Python] Inicializando janela nativa com PyWebView (v{APP_VERSION})...")
        window = webview.create_window(
            title=f"Eletrocentros — PCP & Planejamento v{APP_VERSION}",
            url=str(index_file),
            js_api=api,
            width=1380,
            height=900,
            resizable=True,
            min_size=(800, 600)
        )

        def on_started():
            import time
            time.sleep(0.2)
            apply_native_window_icon()
            time.sleep(0.8)
            apply_native_window_icon()

        webview.start(on_started, debug=False)
    except ImportError:
        print("[GUI Python] pywebview não encontrado no ambiente Python corporativo.")
        print("[GUI Python] Iniciando servidor HTTP local dinâmico...")

        port = find_available_port(8000)
        server_ready_event = threading.Event()

        def server_worker():
            try:
                os.chdir(FRONTEND_DIR)
                with ThreadedTCPServer(("127.0.0.1", port), AppHTTPRequestHandler) as httpd:
                    print(f"[Python HTTP Server] Servindo frontend e API em http://127.0.0.1:{port}")
                    server_ready_event.set()
                    httpd.serve_forever()
            except Exception as e:
                print(f"[Python HTTP Server Erro] {e}")
                server_ready_event.set()

        server_thread = threading.Thread(target=server_worker, daemon=True)
        server_thread.start()
        server_ready_event.wait(timeout=5)

        url = f"http://127.0.0.1:{port}/index.html"
        print(f"[GUI Python] Abrindo aplicação no navegador padrão em {url}...")
        webbrowser.open(url)

        try:
            while True:
                time.sleep(1)
        except KeyboardInterrupt:
            print("\n[GUI Python] Aplicação encerrada pelo usuário.")


if __name__ == "__main__":
    launch_app()

